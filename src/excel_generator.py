import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime, timedelta
from dateutil.parser import parse as parse_date
import json
 
# ── TP Brand Palette ──────────────────────────────────────────────────────────
def fill(c):  return PatternFill("solid", fgColor=c)
 
TP_TEAL_DARK    = fill("0F5160")   # primary header
TP_TEAL_MED     = fill("468389")   # sub-header / accent
TP_TEAL_LIGHT   = fill("6AAFBA")   # column headers
TP_TEAL_ICE     = fill("E4F2F0")   # alt row
TP_ORANGE       = fill("ED8339")   # total / grand total highlight
TP_ORANGE_LIGHT = fill("F4A96A")   # secondary warm
WHITE           = fill("FFFFFF")
GRAY_LIGHTEST   = fill("F0F0F0")
 
def bd_thin():
    s = Side(style="thin", color="B8DDD8")   # teal-lightest border
    return Border(left=s, right=s, top=s, bottom=s)
 
def bd_med():
    s = Side(style="medium", color="468389")
    return Border(left=s, right=s, top=s, bottom=s)
 
def hf(sz=10, col="FFFFFF", bold=True):
    return Font(name="Arial", bold=bold, color=col, size=sz)
 
def bf(sz=10, col="1A1A1A", bold=False):
    return Font(name="Arial", size=sz, color=col, bold=bold)
 
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left",   vertical="center")
RGT = Alignment(horizontal="right",  vertical="center")
 
# ── Week helpers — weeks start from month day 1 ───────────────────────────────
def week_of_month_monday(dt):
    """
    Return the 'week bucket' date = first day of that week-of-month block.
    Weeks are counted from the 1st of the month:
      Day  1–7  → Week 1 (returns month's day 1)
      Day  8–14 → Week 2 (returns month's day 8)
      Day 15–21 → Week 3 (returns month's day 15)
      Day 22–31 → Week 4 (returns month's day 22)
    """
    day = dt.day
    if day <= 7:
        return dt.replace(day=1,  hour=0, minute=0, second=0, microsecond=0)
    elif day <= 14:
        return dt.replace(day=8,  hour=0, minute=0, second=0, microsecond=0)
    elif day <= 21:
        return dt.replace(day=15, hour=0, minute=0, second=0, microsecond=0)
    else:
        return dt.replace(day=22, hour=0, minute=0, second=0, microsecond=0)
 
def week_label(wk):
    """'Apr Wk1 (Apr 01–07)'"""
    month_abbr = wk.strftime("%b")
    day_start  = wk.day
    if day_start == 1:
        wn, day_end = 1, 7
    elif day_start == 8:
        wn, day_end = 2, 14
    elif day_start == 15:
        wn, day_end = 3, 21
    else:
        wn = 4
        # last day of month
        next_m  = (wk.replace(day=28) + timedelta(days=4)).replace(day=1)
        day_end = (next_m - timedelta(days=1)).day
    return f"{month_abbr} Wk{wn}\n({month_abbr} {day_start:02d}–{day_end:02d})"
 
# ── Main builder ──────────────────────────────────────────────────────────────
def build_excel(records, start_dt, end_dt, output_path, project):
    # Parse dates
    for r in records:
        if isinstance(r["started"], str):
            r["started"] = parse_date(r["started"])
 
    # Build pivot: author → week_bucket → hours
    pivot   = defaultdict(lambda: defaultdict(float))
    authors, weeks = set(), set()
 
    for r in records:
        wk = week_of_month_monday(r["started"])
        pivot[r["author"]][wk] += r["hours"]
        authors.add(r["author"])
        weeks.add(wk)
 
    authors = sorted(authors)
    weeks   = sorted(weeks)
 
    label_start = start_dt.strftime("%b %d, %Y")
    label_end   = end_dt.strftime("%b %d, %Y")
 
    wb = openpyxl.Workbook()
 
    # ═══════════════════════════════════════════════════════════════════════════
    # TAB 1 — Raw Worklogs
    # ═══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Raw Worklogs"
 
    # Title
    ws1.row_dimensions[1].height = 30
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    t = ws1.cell(1, 1, f"Transparent Partners  ·  {project} Project  ·  Raw Worklogs")
    t.font = hf(12); t.fill = TP_TEAL_DARK; t.alignment = CTR
 
    # Sub-title
    ws1.row_dimensions[2].height = 16
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    s = ws1.cell(2, 1, f"{label_start}  →  {label_end}   |   {len(records)} entries   |   {len(authors)} contributors")
    s.font = Font(name="Arial", size=9, italic=True, color="FFFFFF")
    s.fill = TP_TEAL_MED; s.alignment = CTR
 
    # Column headers
    ws1.row_dimensions[3].height = 22
    raw_headers = ["#", "Issue Key", "Author", "Date", "Week", "Day of Week", "Hours"]
    for ci, h in enumerate(raw_headers, 1):
        c = ws1.cell(3, ci, h)
        c.font = hf(9); c.fill = TP_TEAL_LIGHT; c.alignment = CTR; c.border = bd_thin()
 
    # Data rows — sorted by date then author
    sorted_records = sorted(records, key=lambda x: (x["started"], x["author"]))
    for ri, r in enumerate(sorted_records, 4):
        rf = TP_TEAL_ICE if ri % 2 == 0 else WHITE
        ws1.row_dimensions[ri].height = 16
        wk = week_of_month_monday(r["started"])
        vals = [
            ri - 3,
            r["issue"],
            r["author"],
            r["started"].strftime("%Y-%m-%d"),
            week_label(wk).replace("\n", " "),
            r["started"].strftime("%A"),
            round(r["hours"], 2),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws1.cell(ri, ci, val)
            c.font = bf(9)
            c.fill = rf
            c.border = bd_thin()
            if ci == 7:   # hours
                c.alignment = RGT
                c.number_format = "#,##0.00"
            elif ci in (4, 5, 6):
                c.alignment = CTR
            else:
                c.alignment = LFT
 
    # Totals row at bottom
    tr1 = 4 + len(sorted_records)
    ws1.row_dimensions[tr1].height = 20
    ws1.cell(tr1, 1, "TOTAL").font = hf(9); ws1.cell(tr1, 1).fill = TP_TEAL_DARK; ws1.cell(tr1, 1).alignment = CTR; ws1.cell(tr1, 1).border = bd_thin()
    for ci in range(2, 7):
        ws1.cell(tr1, ci).fill = TP_TEAL_DARK; ws1.cell(tr1, ci).border = bd_thin()
    # Hours total formula
    c = ws1.cell(tr1, 7, f"=SUM(G4:G{tr1-1})")
    c.font = hf(9); c.fill = TP_ORANGE; c.alignment = RGT
    c.border = bd_thin(); c.number_format = "#,##0.00"
 
    # Column widths
    for col, w in zip("ABCDEFG", [5, 12, 22, 12, 20, 13, 9]):
        ws1.column_dimensions[col].width = w
 
    ws1.freeze_panes = "A4"
 
    # ═══════════════════════════════════════════════════════════════════════════
    # TAB 2 — Weekly Summary by Author
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Weekly Summary")
    nc  = 3 + len(weeks)   # Author | Total | Cumulative | wk1 | wk2 ...
 
    # Title
    ws2.row_dimensions[1].height = 30
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
    t2 = ws2.cell(1, 1, f"Transparent Partners  ·  {project} Project  ·  Weekly Hours by Author")
    t2.font = hf(12); t2.fill = TP_TEAL_DARK; t2.alignment = CTR
 
    # Sub-title
    ws2.row_dimensions[2].height = 16
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
    s2 = ws2.cell(2, 1,
        f"{label_start}  →  {label_end}   |   "
        f"{len(authors)} contributors   |   Weeks counted from month day 1"
    )
    s2.font = Font(name="Arial", size=9, italic=True, color="FFFFFF")
    s2.fill = TP_TEAL_MED; s2.alignment = CTR
 
    # Column headers row
    ws2.row_dimensions[3].height = 40
    fixed_headers = ["Author", "Total Hours", f"Cumulative\n(to {end_dt.strftime('%b %d')})"]
    all_headers   = fixed_headers + [week_label(w) for w in weeks]
    for ci, label in enumerate(all_headers, 1):
        c = ws2.cell(3, ci, label)
        c.font = hf(9)
        c.fill = TP_TEAL_LIGHT if ci > 3 else TP_TEAL_DARK
        c.alignment = CTR
        c.border = bd_thin()
 
    # Data rows
    for ri, author in enumerate(authors, 4):
        ws2.row_dimensions[ri].height = 20
        rf = TP_TEAL_ICE if ri % 2 == 0 else WHITE
 
        # Author name
        c = ws2.cell(ri, 1, author)
        c.font = bf(10, bold=True); c.fill = rf; c.alignment = LFT; c.border = bd_thin()
 
        # Total hours (SUM formula across all week columns)
        wk_first = get_column_letter(4)
        wk_last  = get_column_letter(3 + len(weeks))
        c = ws2.cell(ri, 2, f"=SUM({wk_first}{ri}:{wk_last}{ri})")
        c.font = bf(10, bold=True, col="0F5160")
        c.fill = TP_TEAL_ICE if ri % 2 == 0 else GRAY_LIGHTEST
        c.alignment = RGT; c.border = bd_med(); c.number_format = "#,##0.00"
 
        # Cumulative hours = same as total (all hours up to end_dt)
        c = ws2.cell(ri, 3, f"=B{ri}")
        c.font = bf(10, bold=True, col="ED8339")
        c.fill = fill("FEF3E8")   # warm orange tint
        c.alignment = RGT; c.border = bd_thin(); c.number_format = "#,##0.00"
 
        # Weekly cells
        for ci, wk in enumerate(weeks, 4):
            h = pivot[author].get(wk, 0)
            c = ws2.cell(ri, ci, round(h, 2) if h else 0)
            c.font = bf(10)
            c.fill = rf
            c.alignment = CTR
            c.border = bd_thin()
            c.number_format = '#,##0.00;-#,##0.00;"-"'
 
    # Grand total row
    tr2 = 4 + len(authors)
    ws2.row_dimensions[tr2].height = 24
 
    c = ws2.cell(tr2, 1, "GRAND TOTAL")
    c.font = hf(10); c.fill = TP_TEAL_DARK; c.alignment = LFT; c.border = bd_thin()
 
    c = ws2.cell(tr2, 2, f"=SUM(B4:B{tr2-1})")
    c.font = hf(10); c.fill = TP_ORANGE; c.alignment = RGT
    c.border = bd_thin(); c.number_format = "#,##0.00"
 
    c = ws2.cell(tr2, 3, f"=B{tr2}")
    c.font = hf(10); c.fill = TP_ORANGE; c.alignment = RGT
    c.border = bd_thin(); c.number_format = "#,##0.00"
 
    for ci in range(4, 4 + len(weeks)):
        col_l = get_column_letter(ci)
        c = ws2.cell(tr2, ci, f"=SUM({col_l}4:{col_l}{tr2-1})")
        c.font = hf(10); c.fill = TP_TEAL_MED; c.alignment = CTR
        c.border = bd_thin(); c.number_format = "#,##0.00"
 
    # Column widths
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 16
    for ci in range(4, 4 + len(weeks)):
        ws2.column_dimensions[get_column_letter(ci)].width = 14
    ws2.freeze_panes = "D4"
 
    wb.save(output_path)
    print(f"  ✅  Saved → {output_path}")
 
    # Console summary
    print(f"\n  {'Author':<26} {'Total Hrs':>10}")
    print(f"  {'-'*26} {'-'*10}")
    for a in authors:
        print(f"  {a:<26} {sum(pivot[a].values()):>10.2f}")
    print(f"  {'GRAND TOTAL':<26} {sum(sum(pivot[a].values()) for a in authors):>10.2f}\n")